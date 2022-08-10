from openpyxl import *

def truong(cityID, city, schoolCol, disCol, type):
    wb = load_workbook(r'C:\xampp\htdocs\search_project\config\reconstruct\nguyen-vong.xlsx')
    ws = wb.active
    data_lst = []

    for i in range(3, ws.max_row+1):
        id = cityID.upper() + str(ws[f'A{i}'].value)
        name = ws[f'{schoolCol}{i}'].value
        district = ws[f'{disCol}{i}'].value
        data_lst.append([id, name, city, district, type])
    
    print(data_lst)
    
    wb2 = Workbook()
    ws2 = wb2.active
    ws2.title = "data"

    for i in data_lst:
        ws2.append(i)
    
    wb2.save(r'C:\xampp\htdocs\search_project\config\reconstruct\truong.xlsx')

def diemChuan(cityID):
    wb = load_workbook(r'C:\xampp\htdocs\search_project\config\reconstruct\nguyen-vong.xlsx')
    ws = wb.active

    yearCol = [ ['2015', list('DEF')],
                ['2016', list('GHI')],
                ['2017', list('JKL')],
                ['2018', list('MNO')],
                ['2019', list('PQR')],
                ['2020', list('STU')],
                ['2021', list('VWX')],
                ['2022', 'Y Z AA'.split(' ')]]
    lst = []
    scoreData = [i[1] for i in yearCol]

    year = 2015
    for j in scoreData:
        count = 1
        for k in j:
            for i in range(3, ws.max_row+1):
                # lst.append([cityID + str(ws[f'A{i}'].value), [ws[f'{scores}{i}'].value for scores in j]])
                id = cityID + str(ws[f'A{i}'].value)
                scores = ws[f'{k}{i}'].value
                lst.append([year, id, f'NV{count}', scores])
            count += 1
        year += 1 

    print(lst)

    wb.close()

    wb2 = Workbook()
    ws2 = wb2.active
    ws2.title = "data"

    for i in lst:
        ws2.append(i)

    wb2.save(r'C:\xampp\htdocs\search_project\config\reconstruct\diem_chuan.xlsx')

def chiTieu(cityID):
    wb = load_workbook(r'C:\xampp\htdocs\search_project\config\reconstruct\ti-le-choi.xlsx')
    ws = wb.active

    yearCol = [ ['2015', list('EF')],
                ['2016', list('HI')],
                ['2017', list('KL')],
                ['2018', list('NO')],
                ['2019', list('QR')],
                ['2020', list('TU')],
                ['2021', list('WX')],
                ['2022', 'Z AA'.split(' ')] ]
    lst = []
    nvData = [i[1] for i in yearCol]

    year = 2015
    for j in nvData:
        count = 1
        for k in j:
            for i in range(3, ws.max_row+1):
                id = cityID + str(ws[f'A{i}'].value)
                chitieu = ws[f'{j[0]}{i}'].value
                soluong = ws[f'{j[1]}{i}'].value
                lst.append([year, id, 'NV1', chitieu, soluong])
            count += 1
        year += 1 

    print(lst)

    wb.close()

    wb2 = Workbook()
    ws2 = wb2.active
    ws2.title = "data"

    for i in lst:
        ws2.append(i)

    wb2.save(r'C:\xampp\htdocs\search_project\config\reconstruct\tlc.xlsx')

def diemChuyen():
    wb = load_workbook(r'C:\xampp\htdocs\search_project\config\reconstruct\chuyen.xlsx')
    ws = wb.active

    yearCol = [ ['2015', list('BC')],
                ['2016', list('DE')],
                ['2017', list('FG')],
                ['2018', list('HI')],
                ['2019', list('JK')],
                ['2020', list('LM')],
                ['2021', list('NO')],
                ['2022', list('PQ')]]
    lst = []
    scoreData = [i[1] for i in yearCol]

    for i in range(3, ws.max_row+1):
        for j in scoreData:
            year = int(str(ws[f'{j[0]}1'].value).replace('Năm', ''))
            id = ws[f'A{i}'].value
            nv = ws[f'{j[0]}{i}'].value
            scores = ws[f'{j[1]}{i}'].value
            if (nv != None and scores != None):
                lst.append([year, id, nv, scores])

    print(lst)

    wb.close()

    wb2 = Workbook()
    ws2 = wb2.active
    ws2.title = "data"

    for i in lst:
        ws2.append(i)

    wb2.save(r'C:\xampp\htdocs\search_project\config\reconstruct\diem_chuyen.xlsx')

# truong('HCM_', 'TP Ho Chi Minh', 'B', 'C', 'L02')
# diemChuan('HCM_')
# chiTieu('HCM_')
diemChuyen()

